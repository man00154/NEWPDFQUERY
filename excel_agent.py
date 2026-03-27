import os
import re
import subprocess
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import openpyxl
from rapidfuzz import fuzz, process


# --------------------------
# Helpers: .xls -> .xlsx
# --------------------------
def convert_xls_to_xlsx(xls_path: str, out_dir: str) -> str:
    """
    Converts .xls to .xlsx using LibreOffice (soffice).
    Returns path to converted xlsx.
    """
    os.makedirs(out_dir, exist_ok=True)
    cmd = [
        "soffice",
        "--headless",
        "--convert-to",
        "xlsx",
        "--outdir",
        out_dir,
        xls_path,
    ]
    res = subprocess.run(cmd, capture_output=True, text=True)
    if res.returncode != 0:
        raise RuntimeError(
            f"LibreOffice conversion failed for {xls_path}\nSTDOUT:\n{res.stdout}\nSTDERR:\n{res.stderr}"
        )

    # LibreOffice output file name is usually same base name with .xlsx
    base = os.path.splitext(os.path.basename(xls_path))[0] + ".xlsx"
    out_path = os.path.join(out_dir, base)
    if not os.path.exists(out_path):
        # fallback: find newest xlsx in out_dir
        xlsx_files = [os.path.join(out_dir, f) for f in os.listdir(out_dir) if f.lower().endswith(".xlsx")]
        if not xlsx_files:
            raise FileNotFoundError("Conversion reported success but no .xlsx produced.")
        out_path = max(xlsx_files, key=os.path.getmtime)

    return out_path


def normalize_path_list(paths: List[str], work_dir: str) -> List[str]:
    """
    Ensures all excel files are .xlsx by converting .xls to .xlsx.
    Returns list of usable file paths.
    """
    usable = []
    for p in paths:
        if p.lower().endswith(".xls") and not p.lower().endswith(".xlsx"):
            xlsx = convert_xls_to_xlsx(p, out_dir=work_dir)
            usable.append(xlsx)
        else:
            usable.append(p)
    return usable


# --------------------------
# Raw grid extraction
# (guarantees all cells readable)
# --------------------------
def excel_col_letter(n: int) -> str:
    """1-indexed column number -> Excel letters"""
    letters = ""
    while n:
        n, r = divmod(n - 1, 26)
        letters = chr(65 + r) + letters
    return letters


def read_sheet_raw_grid(xlsx_path: str, sheet_name: str) -> pd.DataFrame:
    """
    Reads every cell in the sheet into a long-form dataframe:
    file, sheet, row, col, a1, value

    This is the most reliable representation for unstructured sheets.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[sheet_name]

    max_row = ws.max_row
    max_col = ws.max_column

    records = []
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str):
                v = v.strip()
                if v == "":
                    v = None
            a1 = f"{excel_col_letter(c)}{r}"
            records.append(
                {
                    "file": os.path.basename(xlsx_path),
                    "sheet": sheet_name,
                    "row": r,
                    "col": c,
                    "a1": a1,
                    "value": v,
                }
            )

    df = pd.DataFrame.from_records(records)
    return df


def raw_grid_to_matrix(raw_df: pd.DataFrame) -> pd.DataFrame:
    """
    Converts long-form raw grid back to a matrix (rows x cols).
    Useful for previewing the sheet as it visually appears.
    """
    if raw_df.empty:
        return pd.DataFrame()

    max_row = int(raw_df["row"].max())
    max_col = int(raw_df["col"].max())
    mat = pd.DataFrame(index=range(1, max_row + 1), columns=range(1, max_col + 1), dtype=object)

    for _, rec in raw_df.iterrows():
        mat.iat[int(rec["row"]) - 1, int(rec["col"]) - 1] = rec["value"]

    mat.index.name = "row"
    mat.columns = [excel_col_letter(c) for c in mat.columns]
    return mat


# --------------------------
# Best-effort table detection
# (for sums/min/max/charts)
# --------------------------
def _score_header_row(values: List[object]) -> float:
    """
    Heuristic score for whether a row looks like a header.
    High score if many strings and few numerics, and not mostly empty.
    """
    non_empty = [v for v in values if v is not None and str(v).strip() != ""]
    if not non_empty:
        return 0.0

    str_cnt = sum(isinstance(v, str) for v in non_empty)
    num_cnt = sum(isinstance(v, (int, float, np.number)) for v in non_empty)

    # Penalize rows that are almost all same value (like repeated section title)
    uniq = len(set(map(lambda x: str(x).strip().lower(), non_empty)))
    diversity = uniq / max(len(non_empty), 1)

    # Encourage "string heavy", discourage "numeric heavy"
    score = (str_cnt / len(non_empty)) * 0.8 + diversity * 0.2 - (num_cnt / len(non_empty)) * 0.5
    return max(score, 0.0)


def detect_table_from_matrix(mat: pd.DataFrame, header_search_rows: int = 30) -> Tuple[pd.DataFrame, Dict]:
    """
    Given a matrix view of a sheet, find a likely header row and return a table DataFrame.
    This is best-effort; raw grid remains the source of truth.

    Returns: (table_df, meta)
    """
    if mat.empty:
        return pd.DataFrame(), {"header_row": None, "confidence": 0.0}

    # Search top N rows for best header candidate
    n = min(header_search_rows, mat.shape[0])
    scores = []
    for r in range(n):
        row_vals = mat.iloc[r, :].tolist()
        scores.append(_score_header_row(row_vals))

    header_idx = int(np.argmax(scores)) if scores else 0
    confidence = float(scores[header_idx]) if scores else 0.0

    header = mat.iloc[header_idx, :].tolist()

    # Clean header names
    def clean_col(x, i):
        if x is None or str(x).strip() == "":
            return f"Unnamed_{i+1}"
        s = re.sub(r"\s+", " ", str(x)).strip()
        return s

    columns = [clean_col(h, i) for i, h in enumerate(header)]

    # Data starts after header row
    data = mat.iloc[header_idx + 1 :, :].copy()
    data.columns = columns

    # Drop fully empty rows
    data = data.dropna(how="all")

    # Best-effort numeric conversion per column
    for col in data.columns:
        data[col] = _smart_to_numeric(data[col])

    meta = {"header_row": int(header_idx + 1), "confidence": confidence}
    return data.reset_index(drop=True), meta


def _smart_to_numeric(series: pd.Series) -> pd.Series:
    """
    Attempts to convert messy numeric strings to floats while keeping non-numeric text unchanged.
    """
    def parse_one(v):
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return np.nan
        if isinstance(v, (int, float, np.number)):
            return float(v)
        s = str(v).strip()

        # Remove commas, currency symbols, stray spaces
        s2 = re.sub(r"[,\$₹€£]", "", s)

        # Percent
        if s2.endswith("%"):
            try:
                return float(s2[:-1].strip()) / 100.0
            except:
                return v

        # Plain number
        try:
            return float(s2)
        except:
            return v

    converted = series.map(parse_one)

    # Keep as numeric only if "mostly numeric"
    numeric_ratio = pd.to_numeric(converted, errors="coerce").notna().mean()
    if numeric_ratio >= 0.6:
        return pd.to_numeric(converted, errors="coerce")
    return converted


# --------------------------
# Search & operations
# --------------------------
def search_raw_cells(raw_df: pd.DataFrame, query: str, limit: int = 200) -> pd.DataFrame:
    """
    Searches for a query across all raw cell values (string contains).
    Returns matching cells with row/col/a1/value.
    """
    if raw_df.empty or not query:
        return pd.DataFrame()

    q = query.strip().lower()

    def match(v):
        if v is None:
            return False
        return q in str(v).lower()

    hits = raw_df[raw_df["value"].map(match)].copy()
    hits = hits.sort_values(["sheet", "row", "col"]).head(limit)
    return hits


def compute_numeric_summary(df: pd.DataFrame) -> pd.DataFrame:
    """
    Numeric summary for detected table view.
    """
    if df.empty:
        return pd.DataFrame()

    num = df.select_dtypes(include=[np.number])
    if num.empty:
        return pd.DataFrame()

    summary = pd.DataFrame({
        "count": num.count(),
        "sum": num.sum(numeric_only=True),
        "min": num.min(numeric_only=True),
        "max": num.max(numeric_only=True),
        "mean": num.mean(numeric_only=True),
        "median": num.median(numeric_only=True),
    })
    return summary.reset_index(names=["column"])


def fuzzy_find_column(df: pd.DataFrame, col_hint: str, threshold: int = 70) -> Optional[str]:
    """
    Fuzzy match a user-provided column hint to a real column name.
    """
    if df.empty or not col_hint:
        return None
    choices = list(df.columns)
    match = process.extractOne(col_hint, choices, scorer=fuzz.WRatio)
    if not match:
        return None
    name, score, _ = match
    return name if score >= threshold else None


def filter_table_by_value(df: pd.DataFrame, column: str, value_substring: str) -> pd.DataFrame:
    """
    Filter rows where df[column] contains substring (case-insensitive).
    """
    if df.empty or column not in df.columns or not value_substring:
        return df

    q = value_substring.strip().lower()

    def contains(v):
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return False
        return q in str(v).lower()

    return df[df[column].map(contains)].copy()


@dataclass
class WorkbookData:
    path: str
    sheets: List[str]
    raw_by_sheet: Dict[str, pd.DataFrame]          # long-form raw grids
    matrix_by_sheet: Dict[str, pd.DataFrame]       # matrix view (for preview)
    table_by_sheet: Dict[str, pd.DataFrame]        # detected table view
    table_meta_by_sheet: Dict[str, Dict]           # detection meta


def load_workbook_all_views(xlsx_path: str) -> WorkbookData:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    sheets = wb.sheetnames

    raw_by_sheet = {}
    matrix_by_sheet = {}
    table_by_sheet = {}
    table_meta_by_sheet = {}

    for s in sheets:
        raw = read_sheet_raw_grid(xlsx_path, s)
        mat = raw_grid_to_matrix(raw)
        table, meta = detect_table_from_matrix(mat)

        raw_by_sheet[s] = raw
        matrix_by_sheet[s] = mat
        table_by_sheet[s] = table
        table_meta_by_sheet[s] = meta

    return WorkbookData(
        path=xlsx_path,
        sheets=sheets,
        raw_by_sheet=raw_by_sheet,
        matrix_by_sheet=matrix_by_sheet,
        table_by_sheet=table_by_sheet,
        table_meta_by_sheet=table_meta_by_sheet,
    )
